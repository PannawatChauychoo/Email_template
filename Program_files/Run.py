import pandas as pd 
import time 
from openpyxl import Workbook, load_workbook
import datetime as dt

from Check_creds import generate_credentials    #Authentication of user
from datetime_conversion import convert_str, custom_strftime, convert_dt    #Formatting datetime into str
from Generate_templates import write_email    #Return draft strings in html code
from Generate_draft_pretty import create_htmldraft    #Create email drafts with rich formatting
from Check_time_record import check_ideal_sendtime, record_followup_time, update_status    #Update excel file with the time
from Send_draft import gmail_send_draft, check_if_sent    #Sending email draft and check if sent 


"""
Goal of functionality:
- Generate email template based on the time we have reached out
- Check current date and output the ideal date
- Create draft and send email automatically
- Record the date when creating the draft and when the email is sent
- Record the time when we should reach out again 
- Automatically update the cells with appropriate values 
- Bonus: Run at least once a day in the background (schedule + nohup ... &) - https://stackoverflow.com/questions/15088037/python-script-to-do-something-at-the-same-time-every-day

"""

def row_into_string(row_list):
    #Combining all values of a row into a string
    output_string = ", ".join([str(i) for i in row_list])
    return output_string

def main():
    #Getting user credentials
    generate_credentials()
    
    data_path = '../Data/Contact_tracking.xlsx'
    target = pd.read_excel('../Data/Contact_tracking.xlsx', sheet_name= "Targets")
    my_info = pd.read_excel('../Data/Contact_tracking.xlsx', sheet_name= "Myself")
    
    
    #Combining my info with availabilities into a string
    my_email = my_info.iloc[0, 0]
    my_info_strings = my_info.iloc[:, 1:6].apply(row_into_string, axis = 1)
    avail = my_info.iloc[0, 6:]

    my_formatted_info = my_info_strings + ", " + ", ".join([custom_strftime(i) for i in avail]) #Getting the right format for dates

    #Combining all columns into strings per row
    target_input_strings = target.iloc[:, 1:7].apply(row_into_string, axis = 1) 
    
    #Create and format the ideal sent time, current time and next reach out time 
    ideal_time = check_ideal_sendtime()
    ideal_time_formatted = convert_dt(ideal_time)
    
    time_now = dt.datetime.now()
    time_now_formatted = convert_dt(time_now)
    
    next_reach_out = ideal_time + dt.timedelta(days = 5)
    next_reach_out_formatted = convert_dt(next_reach_out)
    
    #Opening the excel file 
    # 
    # wb = load_workbook(data_path)
    # target_sheet = wb.active
    
    #Create a draft for every email
    for i in range(0, len(target)):
        target_email = target.iloc[i, 0] #Target email
        follow_up_count = target.iloc[i, 6] 
        response = target.iloc[i, 7] #Responded or not 
        indexed_row = str(i + 2) #Skip the first header row 
        status_check = target.iloc[i, 14] #Getting the status
        
        #Check if they have responded
        if response.lower() == 'yes':
            print('Awesome! Goodluck with the informational interview.')
        elif response.lower() == 'no':
            if follow_up_count == 0: #First time 
                try:
                    #generate template
                    email_draft_template = write_email(target_input_strings[i], my_formatted_info[0])
                    draft = create_htmldraft(email_draft_template, target_email)
                    draft_id = draft["id"]
                    print(f'Draft for {target_email} created with id: {draft_id}')
                    
                    #Update excel files with the new follow_up_count and sent out time
                    record_followup_time(indexed_row, follow_up_count, draft_id, ideal_time_formatted, time_now_formatted, next_reach_out_formatted)
                    #Update follow_up_count, update status of email & input draft_id 
                    update_status(indexed_row, time_now, ideal_time)
                    
                except Exception as error:
                    print(error)

            elif follow_up_count < 3:   #Second & third follow up
                try:
                    wb = load_workbook(data_path)
                    target_sheet = wb.active
                    scheduled_time_cell = 'N' + indexed_row
                    scheduled_time = convert_str(target_sheet[scheduled_time_cell].value)
                    update_status(indexed_row, time_now, scheduled_time)
                    wb.save(data_path)
                    
                    if status_check == 'Pending':    #Email waiting to be sent - no action
                        pass
                    elif status_check == 'Sending':    #Sending emails
                        draft_id = target.iloc[i, 12]
                        send_time = convert_str(target.iloc[i, 13])
                        sent_email = gmail_send_draft(draft_id = draft_id, recipients = target_email)
                        
                        wb = load_workbook(data_path)
                        target_sheet = wb.active
                        #Update status
                        status_cell = 'O' + indexed_row
                        target_sheet[status_cell].value = "Sent"
                        
                        #Adding message ID and thread ID for future reference
                        message_ID_cell = 'Q' + indexed_row
                        target_sheet[message_ID_cell].value = sent_email["id"]
                        thread_ID_cell = 'R' + indexed_row
                        target_sheet[thread_ID_cell].value = sent_email["threadId"]
                                
                        print(f"Row {indexed_row}: Email for {target_email} has been sent!")
                        wb.save(data_path)
                        
                    elif status_check == 'Sent?':    #Checking if an email should have been sent 
                        message_ID = target.iloc[i, 16]
                        sent_code = check_if_sent(message_id=message_ID) 
                        
                        wb = load_workbook(data_path)
                        target_sheet = wb.active
                        if sent_code == 1:
                            print(f"Row {indexed_row}: Email has been sent!")
                            status_cell = 'O' + indexed_row
                            target_sheet[status_cell].value = "Sent"
                        else:
                            print(f"Row {indexed_row}: Email for {target_email} was not sent!")
                            pass
                        wb.save(data_path)    
                    
                    elif status_check == 'Sent': 
                        Next_follow_up_date = convert_str(target.iloc[i, 15])
                        if time_now < Next_follow_up_date:
                            pass
                        elif time_now >= Next_follow_up_date:
                            email_draft_template = write_email(target_input_strings[i], my_formatted_info[0])
                            draft = create_htmldraft(email_draft_template, email)
                            draft_id = draft["id"]
                            print(f"Row {indexed_row}: Draft for {target_email} created with id: {draft_id}")
                            record_followup_time(indexed_row, follow_up_count, draft_id, ideal_time_formatted, time_now_formatted, next_reach_out_formatted)
                            update_status(indexed_row, time_now, ideal_time)
                        else: 
                            print("Something is missing")
                            
                except Exception as error:
                    print(error)
            elif follow_up_count == 3:
                pass
            else:
                print(f"Row {indexed_row} has out of index follow up count. Moving onto others.")
                pass
    # wb.save(data_path)
    
if __name__ == "__main__":
    main()