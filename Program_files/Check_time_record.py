import datetime as dt 
import random
import time

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

#Tuesday -> Thursday from 8-10am are go time
def check_ideal_sendtime(prefer_weekdays = [1,2,3], prefer_hours = [8,9,10]):
    """
    Finding the ideal time to send based on current time
    """
    now = dt.datetime.now() #example: datetime.datetime(2024, 3, 6, 20, 56, 59, 727333)
    go_days = prefer_weekdays
    go_hours = [8,9,10]

    year = now.year
    weekday = now.weekday() #int (0-6) -> 0 = Monday, 6 = Sunday
    hour = now.hour 
    date = now.day
    month = now.month

    random_hour = random.randint(8,10)
    random_minute = random.randint(0, 59)

    today_date_str = now.strftime("%A") #weekday str
    
    if weekday in go_days:
        if hour in go_hours:
            print(f"Today is {today_date_str} and in prime_time! Let's send them!")
            scheduled_time = now
        elif hour not in go_hours and weekday < max(go_days):
            print(f"Today is {today_date_str} but not in primetime! Let's send them tomorrow!")
            scheduled_time = dt.datetime(year, month, date + 1, random_hour, random_minute)
        elif hour not in go_hours and weekday >= max(go_days):
            wait_time = 6 - weekday + min(go_days) # Count the days to next Tuesday
            scheduled_time = dt.datetime(year, month, date + wait_time, random_hour, random_minute)
            scheduled_day = scheduled_time.strftime("%A")
            print(f"Today is {today_date_str} but we missed the ideal time. Let's schedule emails for next week on {scheduled_day}")        
    elif weekday > max(go_days):
        wait_time = 7 - weekday + min(go_days) # Count the days to next Tuesday
        scheduled_time = dt.datetime(year, month, date + wait_time, random_hour, random_minute)
        scheduled_day = scheduled_time.strftime("%A")
        print(f"Today is {today_date_str}. Let's schedule emails for next week on {scheduled_day}")
    elif weekday < min(go_days):
        scheduled_time = dt.datetime(year, month, date + 1, random_hour, random_minute) #Since it is Monday then schedule on Tuesday
        print(f"Today is {today_date_str}. Let's schedule emails for tomorrow!")
    else: 
        pass
    
    return scheduled_time

def update_status(row, time_now, scheduled_time, excel_name = '../Data/Contact_tracking.xlsx'):
    """
    Checking sent status and update the right column
    Update update status of email & input draft_id
    Return: 0-not send, 1-sending
    """
    wb = load_workbook(excel_name)
    target_sheet = wb['Targets']
    
    row = str(row)
    
    #Check sent status by comparing
    status_cell = 'O' + row
    
    if time_now.day < scheduled_time.day:
        target_sheet[status_cell].value = "Pending"
    elif time_now.day == scheduled_time.day:
        if time_now.hour <= scheduled_time.hour:
            target_sheet[status_cell].value = "Sending"
        elif time_now.hour > scheduled_time.hour:
            target_sheet[status_cell].value = "Sent?"
    else:
        target_sheet[status_cell].value = "Done"
    
    return wb.save(excel_name)  

    
 
def record_followup_time(row, follow_up_count, draft_id, format_scheduled_time, format_time_now, next_reach_out, excel_name = '../Data/Contact_tracking.xlsx'):
    """
    Input: row, follow_up_count, scheduled_time, excel_filename (default = Contact_tracking)
    Insert the time when the email is being sent out under the right follow up count
    Output: saved excel notebook with updated cells
    """
    wb = load_workbook(excel_name)
    target_sheet = wb['Targets']
    row = str(row)
    
    #Inputting the current time into the right follow up count 
    count_dict = {0: 'I', 1:'J', 2:'K'}
    time_column = count_dict[follow_up_count] #Picking the right column based on follow_up_count
    row = row
    cell = str(time_column) + row
    target_sheet[cell].value = format_scheduled_time
    
    #update the follow up count (only if less than 3)
    followup_cell = 'G' + row
    target_sheet[followup_cell].value = follow_up_count + 1 

    #update time created draft 
    draft_cell = 'L' + row
    target_sheet[draft_cell].value = format_time_now
    
    #Input draft ID
    draftid_cell = 'M' + row
    target_sheet[draftid_cell].value = draft_id
    
    #update scheduled sent time
    scheduled_sent_cell = 'N' + row
    target_sheet[scheduled_sent_cell].value = format_scheduled_time
    
    #update next reach out time
    next_reachout_cell = 'P' + row
    target_sheet[next_reachout_cell].value = next_reach_out
    
    wb.save(excel_name)
